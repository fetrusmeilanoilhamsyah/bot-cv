import os
import re
import csv
import logging
import asyncio
from io import BytesIO
from telegram import Update, Document
from telegram.ext import ContextTypes
from database import db
from middleware.session import get_user_dir, clear_user_dir
from core.utils import sanitize_filename

logger = logging.getLogger(__name__)

STATE = "XLSX2TXT_COLLECTING"

# Regex cerdas untuk menangkap format Internasional (+593, 1, 60, dll) maupun Lokal (08xx)
# Didukung penangkapan spasi, strip, atau kurung (misal: +593 99-341-1006)
PHONE_REGEX = re.compile(r'\+?(?:\d[\s\-\(\)\.]*){8,16}')

def _extract_numbers_sync(filepath: str, ext: str) -> list:
    """Ekstrak nomor HP dari Excel/CSV secara sinkron via openpyxl/csv (BERURUTAN)"""
    numbers = []
    seen = set()
    try:
        def process_cell(cell_value):
            if not cell_value: return
            text = str(cell_value)
            for m in PHONE_REGEX.findall(text):
                clean_num = re.sub(r'[^0-9]', '', m)
                if clean_num.startswith("08"):
                    clean_num = "62" + clean_num[1:]
                if 8 <= len(clean_num) <= 15 and clean_num not in seen:
                    seen.add(clean_num)
                    numbers.append(clean_num)

        if ext == ".csv":
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    for cell in row:
                        process_cell(cell)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        process_cell(cell)
            wb.close()
            
    except Exception as e:
        logger.error("Error ekstrak %s: %s", filepath, e)
    return numbers

async def cmd_xlsxtotxt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from middleware.auth import require_member
    if not await require_member(update, context):
        return
        
    user_id = update.effective_user.id
    db.increment_usage(user_id)
    clear_user_dir(user_id)
    
    user_dir = get_user_dir(user_id)
    master_txt = os.path.join(user_dir, "extracted_numbers.txt")
    open(master_txt, 'w').close()
    
    db.set_session(user_id, STATE, {"total_kontak": 0, "total_file": 0})
    await update.message.reply_text(
        "Kirim file Excel (.xlsx) atau CSV (.csv).\n"
        "Klik /done jika selesai."
    )

async def handle_xlsxtotxt_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != STATE:
        return
        
    doc = update.message.document
    if not doc or not doc.file_name:
        await update.message.reply_text("Kirim file .xlsx atau .csv.")
        return
        
    ext = os.path.splitext(doc.file_name)[1].lower()
    if ext not in [".xlsx", ".xls", ".csv"]:
        await update.message.reply_text("Format tidak didukung.")
        return
        
    user_dir = get_user_dir(user_id)
    file_path = os.path.join(user_dir, doc.file_name)
    
    base_name, ex = os.path.splitext(doc.file_name)
    counter = 1
    while os.path.exists(file_path):
        file_path = os.path.join(user_dir, f"{base_name}_{counter}{ex}")
        counter += 1
        
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(file_path)
    except Exception as e:
        logger.error("Download error user %s: %s", user_id, e)
        await update.message.reply_text("Gagal mengunduh file. Coba kirim ulang.")
        return

    loop = asyncio.get_running_loop()
    found_numbers = await loop.run_in_executor(None, _extract_numbers_sync, file_path, ext)
    
    master_txt = os.path.join(user_dir, "extracted_numbers.txt")
    try:
        with open(master_txt, 'r', encoding='utf-8') as f:
            existing = f.read().splitlines()
        
        # Gabungkan dengan tetap menjaga urutan & buang duplikat
        seen = set(existing)
        combined = list(existing)
        for num in found_numbers:
            if num not in seen:
                seen.add(num)
                combined.append(num)
        
        with open(master_txt, 'w', encoding='utf-8') as f:
            f.write("\n".join(combined))
            
        new_total = len(combined)
    except Exception as e:
        logger.error("Error nulis hasil: %s", e)
        new_total = 0

    data = sess["data"]
    data["total_file"] += 1
    data["total_kontak"] = new_total
    db.set_session(user_id, STATE, data)
    
    await update.message.reply_text(
        f"{len(found_numbers)} kontak unik di {doc.file_name}.\n"
        f"Total: {new_total} ({data['total_file']} file)."
    )

async def handle_xlsxtotxt_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != STATE:
        return
        
    data = sess["data"]
    total = data.get("total_kontak", 0)
    
    if total == 0:
        await update.message.reply_text("Nomor tidak ditemukan.")
        db.clear_session(user_id)
        clear_user_dir(user_id)
        return
        
    user_dir = get_user_dir(user_id)
    master_txt = os.path.join(user_dir, "extracted_numbers.txt")
    
    try:
        with open(master_txt, 'rb') as f:
            buffer = BytesIO(f.read())
            buffer.name = "Hasil_Ekstrak_Excel.txt"
            
        await update.message.reply_document(
            document=buffer,
            caption=f"Selesai. {total} kontak unik."
        )
    except Exception as e:
        logger.error("Error kirim hasil xlsx: %s", e)
        await update.message.reply_text("Gagal mengirim hasil. Coba ulangi.")
        
    finally:
        db.clear_session(user_id)
        clear_user_dir(user_id)
